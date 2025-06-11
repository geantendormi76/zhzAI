import os
import openpyxl
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

SAMPLE_DOCS_DIR = "sample_docs"

def analyze_xlsx_with_openpyxl(file_path):
    logger.info(f"\n--- Analyzing XLSX with openpyxl: {file_path} ---")
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True) # data_only=True 获取单元格计算后的值而非公式
        logger.info(f"  Workbook '{os.path.basename(file_path)}' loaded. Sheet names: {workbook.sheetnames}")

        for sheet_name in workbook.sheetnames:
            logger.info(f"\n  --- Processing Sheet: {sheet_name} ---")
            sheet = workbook[sheet_name]
            
            # 策略1: 将整个工作表转换为Markdown格式的文本 (简化版)
            markdown_text = f"Sheet: {sheet_name}\n"
            header_processed = False
            max_cols = sheet.max_column
            
            # 尝试构建Markdown表格
            # 注意：openpyxl不直接知道表格的实际边界，它会迭代所有有数据的单元格
            # 对于稀疏表格或多个表格在同一sheet的情况，这可能不完美
            
            # 简单的Markdown表头
            # header_row_values = [str(sheet.cell(row=1, column=col_idx+1).value or "") for col_idx in range(max_cols)]
            # if any(h.strip() for h in header_row_values): # 只有当表头不全为空时才添加
            #     markdown_text += "| " + " | ".join(header_row_values) + " |\n"
            #     markdown_text += "| " + " | ".join(["---"] * len(header_row_values)) + " |\n"
            #     header_processed = True

            # 逐行转换为Markdown或描述性文本
            extracted_rows_data = []
            for row_idx, row in enumerate(sheet.iter_rows()):
                row_values = [str(cell.value or "").strip() for cell in row[:max_cols]] # 只取有数据的列
                
                # 跳过完全空的行
                if not any(row_values):
                    continue
                    
                extracted_rows_data.append(row_values)
                
                # 转换为Markdown行 (如果选择此策略)
                if row_idx == 0 and not header_processed: # 如果是第一行且没有处理过表头
                    markdown_text += "| " + " | ".join(row_values) + " |\n"
                    markdown_text += "| " + " | ".join(["---"] * len(row_values)) + " |\n"
                else:
                    markdown_text += "| " + " | ".join(row_values) + " |\n"

            logger.info("\n  [Content as Markdown (Simplified)]")
            logger.info(markdown_text)

            # 策略2: 每行转换为描述性文本
            descriptive_text = f"Sheet: {sheet_name}\n"
            if extracted_rows_data:
                headers = extracted_rows_data[0] # 假设第一行为表头
                for row_idx, data_row in enumerate(extracted_rows_data[1:], start=1):
                    descriptive_text += f"Row {row_idx}: "
                    parts = []
                    for col_idx, cell_value in enumerate(data_row):
                        if col_idx < len(headers):
                            parts.append(f"'{headers[col_idx]}' is '{cell_value}'")
                        else: # 如果数据行比表头长
                            parts.append(f"'Column {col_idx+1}' is '{cell_value}'")
                    descriptive_text += "; ".join(parts) + ".\n"
            
            logger.info("\n  [Content as Descriptive Text (per row)]")
            logger.info(descriptive_text)

    except Exception as e:
        logger.error(f"Error processing file {file_path} with openpyxl: {e}", exc_info=True)

if __name__ == "__main__":
    if not os.path.isdir(SAMPLE_DOCS_DIR):
        logger.error(f"Sample documents directory '{SAMPLE_DOCS_DIR}' not found.")
    else:
        xlsx_files = sorted([f for f in os.listdir(SAMPLE_DOCS_DIR) if f.lower().endswith(".xlsx") and not f.startswith("~$")])
        if not xlsx_files:
            logger.warning(f"No .xlsx files found in '{SAMPLE_DOCS_DIR}'.")
        else:
            for filename in xlsx_files:
                file_path = os.path.join(SAMPLE_DOCS_DIR, filename)
                analyze_xlsx_with_openpyxl(file_path)